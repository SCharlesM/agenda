"""
Agenda is a program to make writing
an agenda easier. Input agenda item names and
times and have an easy way to change times
"""
from datetime import datetime, time, timedelta, date
from openpyxl import Workbook, load_workbook
import string
#import csv.py

class Agenda:
    def __init__(self, name, startTime):
        self.name = name

        ##take a string as input for the time
        self.startingTime = startTime

        ##convert into a time object to keep track of current_time
        self.time_object_current = datetime.strptime(self.startingTime, "%H:%M:%S")
        self.string_current_time = str(self.time_object_current.time())

        #initialise an empty list and add the header info
        self.agendaList = []
        header_list = ['Index', 'Start-time', 'Endtime', 'Topic', 'Duration']
        self.agendaList.append(header_list)

    def addAgendaItem(self, name, duration):
        """
        a function to add an AgendaItem,
        a list of 5 variables: Index, Starttime, Endtime, Topic, duration
        """        
        agendaItemList = []
        index = len(self.agendaList)
        agendaItemList.append(index+1)

        #initialise the start time from the Agenda class current time
        #current_time = str(self.string_current_time)
        agendaItemList.append(str(self.time_object_current.time()))

        #add the duration to get 'endtime' and update 'current_time'
        duration_to_add = int(duration)
        time_object_end = self.time_object_current + timedelta(minutes = int(duration_to_add))
        self.time_object_current = time_object_end

        agendaItemList.append(str(time_object_end.time()))
        agendaItemList.append(name)
        agendaItemList.append(duration)
       
        #add this list to the main agenda list
        self.agendaList.append(agendaItemList)

    """
    a print function to print a header and iterate through the
    agendaList to print each item
    """
    def printAgenda(self):
                
        #print a header
        print("--------------------------------")
        print("Title: " + self.name)
        print("Start time: " + self.startingTime)
        print("--------------------------------")
        """
        s1 = "Index"
        s2 = "Startime"
        s3 = "Endtime"
        s4 = "Topic"
        s5 = "Duration"
        print( '{:<10s} {:<10s} {:<10s} {:<15s} {:<10s} '.format(s1, s2, s3, s4, s5))
        #print(self.agendaList)
        """
        #iterate throught the agenaList and print each individual list
        for j in range(0, len(self.agendaList)):

            new_list = self.agendaList[j]

            string_index = str(new_list[0])
            string_starttime = new_list[1]
            string_endtime = new_list[2]
            string_topic = new_list[3]
            string_duration = new_list[4]
        
            #format the strings, left align 20, left align 10 (string)
            print( '{:<10s} {:<10s} {:<10s} {:<15s} {:<10s}'.format(string_index, string_starttime, string_endtime, string_topic, string_duration))

    """
    a function to change title of an AgendaItem, using the index of that AgendaItem
    """
    def editAgendaItemTitle(self, index, new_title):

        temp_list = self.agendaList[index-1]
        temp_list[3] = new_title

    """
    a function to change duration of an AgendaItem, using the index of that item
    """
    def editAgendaItemDuration(self, index, new_duration):

        temp_list = self.agendaList[index-1]
        temp_list[4] = new_duration

    """
        A function to export the Agenda to and Excel document
    """
    def exportToExcel(self, filename):

        #initialist the workbook
        doc_file_name = str(filename)
        workbook = Workbook()
        sheet = workbook.active

        #add data to the workbook
        #sheet["A1"] = str(self.five_agenda_list)

        new_list = self.agendaList[0]

        #populate a list with the uppercase letters
        uppercase_list = []
        for j in string.ascii_uppercase:
            uppercase_list.append(j)

        for l in range(0 , len(self.agendaList)):

            new_list = self.agendaList[l]

            for k in range(0, 5):
                temp_string = (uppercase_list[k] + str(l+1))
                sheet[temp_string] = new_list[k]

        #save to the workbook using the filename argument
        workbook.save(filename=doc_file_name)

        #confirm workbook has been saved
        print("\nData has been exported and saved with filename: " + doc_file_name)

    def inputFromExcel(self, filename):

        Workbook = load_workbook(filename='agenda_input.xlsx')
"""
main function to take input and control output
"""
"""
def main():
    print("Enter your name")
    x = input()
    print("Hello " + x)

if __name__ == "__main__":
    main()        
"""