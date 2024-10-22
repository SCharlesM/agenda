"""
Agenda is a simple program to make writing an Agenda easier

agenda.py takes 'agenda_input.xlsx' excel file as an input that includes a Title, 
start time and a list of topics and durations. agenda.py converts this into an agenda
by adding a individual start and end time for each topic based on the duration of each
topic. It outputs the resulting agenda to "agenda_output<timestamp>.xlsx"
"""

from datetime import datetime, timedelta
from time import strftime, localtime
import string
from openpyxl import Workbook, load_workbook

class Agenda:

    """
    initialise a new agenda with initial variables
    agenda_name is a string
    agenda_starttime is an empty time object
    agenda_list will be used to hold the dictionarys with each agenda item data
    """
    def __init__(self):
        self.agenda_name = ""
        self.agenda_starttime = "00:00:00"
        self.agenda_list = []
        self.time_object_current = datetime.strptime(self.agenda_starttime, "%H:%M:%S")


    def populate_agenda(self, excel_data) :
        """Function to populate agenda list from excel input, stored in a dictionary"""

        for index, entry in enumerate(excel_data, start=1) :
         
            #extract the tuples from the excel_input and initialise the dictionary
            topic_title, topic_duration = entry
            agenda_dict = {
                'index': str(index), 
                'start': "",
                'end': "", 
                'topic': topic_title,
                'duration': str(topic_duration)
            }

            #add the start and end times using the time objects
            agenda_dict["start"] = str(self.time_object_current.time())
            self.time_object_current += timedelta(minutes = topic_duration)
            agenda_dict["end"] = str(self.time_object_current.time())

            self.agenda_list.append(agenda_dict)
            #index += 1

    def print_agenda(self):
        """Function to print a header and iterate through the agenda list to print each item"""

        print("----------------------------------------------------------")
        print("Title: ", self.agenda_name, "\nStart time: ", self.agenda_starttime)
        print("----------------------------------------------------------")
        header = f"{'Index':<10} {'Start':<10} {'End':<11}{'Title':<16}{'Duration':<10}"
        print(header)

        for entry in self.agenda_list :
            s1, s2, s3, s4, s5 = entry.values()
            agenda_entries = f"{s1:<10} {s2:<10} {s3:<10} {s4:<16} {s5:<10}"
            print(agenda_entries)

    def export_to_excel(self):
        """Function to export the Agenda to an Excel document"""
        workbook = Workbook()
        sheet = workbook.active

        #change the whitespace in agenda_name to '_' and use name and timestamp as output filename
        agenda_name_no_space = self.agenda_name.replace(" ", "_")
        timestamp = strftime("_%d_%m_%Y_%I.%M%p", localtime())
        doc_file_name = agenda_name_no_space + timestamp + '.xlsx'

        #set the Title and Start-time at the top of the sheet
        sheet['A1'] = 'Title:'
        sheet['B1'] = self.agenda_name
        sheet['A2'] = 'Start-time:'
        sheet['B2'] = self.agenda_starttime
        sheet['A4'] = 'Index'                       #can this be done any better using the keys?
        sheet['B4'] = 'Start'
        sheet['C4'] = 'End'
        sheet['D4'] = 'Topic'
        sheet['E4'] = 'Duration'

        #iterate though the excel cells to store the agenda item data into each cell
        #item is the dictionary, need to unpack the values and store in excel
        i = 5
        for item in self.agenda_list:
            j = 1
            for key in item :
                cell_reference = sheet.cell(row=i, column=j)
                cell_reference.value = item[key]
                j += 1
            i += 1

        #save to the workbook using the filename argument.
        file_path = 'C:\\Users\\Steve\\Documents\\Coding\\python\\agenda_project\\agenda\\outputs\\'
        workbook.save(file_path + doc_file_name)

        print("\nData has been exported and saved with filename: " + doc_file_name)

    def input_from_excel(self, filename):
        """Function to extract data from input.xlsx"""
        value_list = []

        try :
            workbook = load_workbook(filename)
            sheet = workbook.active

            #set the title and startime of the agenda
            self.agenda_name = sheet["B2"].value
            self.agenda_starttime = str(sheet["B3"].value)

            #iterate through the rows and add the tuple of values to the list and return the list
            #need to add something about checking how long the input is and changing max_rows
            for row in sheet.iter_rows(min_row = 5, max_row = 14, min_col = 1, max_col = 2, values_only = True):

                value_list.append(row)

        except FileNotFoundError:
            print("\nThe file was not found, please check the path and try again")

        return value_list

if __name__ =="__main__":

    agenda1 = Agenda()

    #populate a list with session titles and durations from the excel input file
    excel_input = agenda1.input_from_excel("agenda_input.xlsx")

    if not excel_input :
        print("\nagenda_input.xlsx is empty or doesn't exist, please check the file and try again")
    else :

        #format the starttime date.time object
        agenda1.time_object_current = datetime.strptime(agenda1.agenda_starttime, "%H:%M:%S")

        #populate the agenda_list with the dictionaries containing all the agenda items
        agenda1.populate_agenda(excel_input)

        #print the agenda to the commandline but also export to excel
        agenda1.print_agenda()
        agenda1.export_to_excel()
